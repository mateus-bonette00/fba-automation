import sys
import os
from openpyxl import load_workbook

files_to_fix = [
    "exports/ARQUIVOS XLSX/fcpeuro.com_part_001.xlsx",
    "exports/ARQUIVOS XLSX/fcpeuro.com_part_002.xlsx",
    "exports/ARQUIVOS XLSX/fcpeuro.com_part_003.xlsx",
    "exports/ARQUIVOS XLSX/hartvillehardware.com_part_001.xlsx"
]

base_dir = "/home/mateus/Documentos/Qota Store/códigos/fba-automation"

style_block = """<style>
            :root { --bg: #ffffff; --text: #333333; --table-bg: #f9f9f9; --border: #dddddd; --th-bg: #0563C1; --th-text: #ffffff; --link: #0066cc; --link-hover: #003399; --link-visited: #cc0000; --tr-hover: #f1f1f1; --clicked-bg: #ffe6e6; }
            @media (prefers-color-scheme: dark) {
              :root { --bg: #121212; --text: #e0e0e0; --table-bg: #1e1e1e; --border: #333333; --th-bg: #0563C1; --th-text: #ffffff; --link: #4da6ff; --link-hover: #80bfff; --link-visited: #ff8080; --tr-hover: #2a2a2a; --clicked-bg: #4a1515; }
            }
            body { font-family: monospace; font-size: 14px; margin: 20px; background: var(--bg); color: var(--text); }
            table { border-collapse: collapse; width: 100%; margin-top: 10px; background: var(--table-bg); }
            th, td { border: 1px solid var(--border); padding: 8px; text-align: left; transition: background-color 0.3s; }
            th { background-color: var(--th-bg); color: var(--th-text); position: sticky; top: 0; }
            a { color: var(--link); text-decoration: none; display: block; width: 100%; height: 100%; }
            a:hover { text-decoration: underline; color: var(--link-hover); }
            a:visited { color: var(--link-visited); }
            tr:hover { background-color: var(--tr-hover); }
            td.clicked-cell { background-color: var(--clicked-bg) !important; }
</style>"""

js_script = """
        <script>
            document.addEventListener('DOMContentLoaded', () => {
                const links = document.querySelectorAll('td a');
                
                // Load clicked state from localStorage
                links.forEach(link => {
                    const url = link.getAttribute('href');
                    if(url && localStorage.getItem('clicked_' + url)) {
                        link.parentElement.classList.add('clicked-cell');
                    }
                    
                    // Add click listener
                    link.addEventListener('click', function() {
                        if(url) {
                            localStorage.setItem('clicked_' + url, 'true');
                            this.parentElement.classList.add('clicked-cell');
                        }
                    });
                });
            });
        </script>
"""

for filename in files_to_fix:
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        print(f"File not found: {path} (Certifique-se que o caminho está correto)")
        continue
        
    print(f"Gerando HTML Premium para {filename}...")
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        html_path = path.replace(".xlsx", ".html")
        html_content = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Produtos Capturados</title>",
            style_block,
            "</head><body>",
            f"<h2>Lote de Produtos Exportados: {os.path.basename(filename)}</h2>",
            "<p>Dica: Segure <b>CTRL + Clique</b> nos links abaixo para abrir na hora sem sair dessa tela!</p>",
            "<table><thead><tr><th>Indice</th><th>Produto</th><th>UPC</th><th>URL Fornecedor</th><th>Amazon UPC</th><th>Amazon Título</th></tr></thead><tbody>"
        ]
        
        index = 1
        for row in range(2, ws.max_row + 1):
            produto = ws.cell(row=row, column=1).value or ""
            upc = ws.cell(row=row, column=2).value or ""
            
            # Fetch URLs
            def get_url(cell):
                if cell.value and isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
                    try:
                        return cell.value.split('"')[1]
                    except: return ""
                if cell.hyperlink and cell.hyperlink.target:
                    return cell.hyperlink.target
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                    return cell.value
                return ""
            
            url = get_url(ws.cell(row=row, column=3))
            amz_upc = get_url(ws.cell(row=row, column=4))
            amz_tit = get_url(ws.cell(row=row, column=5))

            html_content.append("<tr>")
            html_content.append(f"<td>{index}</td>")
            html_content.append(f"<td>{produto}</td>")
            html_content.append(f"<td>{upc}</td>")
            html_content.append(f"<td><a href='{url}' target='_blank'>Abrir Fornecedor</a></td>" if url else "<td></td>")
            html_content.append(f"<td><a href='{amz_upc}' target='_blank'>Buscar UPC na Amazon</a></td>" if amz_upc else "<td></td>")
            html_content.append(f"<td><a href='{amz_tit}' target='_blank'>Buscar Titulo na Amazon</a></td>" if amz_tit else "<td></td>")
            html_content.append("</tr>")
            index += 1
            
        html_content.append("</tbody></table>")
        html_content.append(js_script)
        html_content.append("</body></html>")
        
        with open(html_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html_content))
            
        print(f"Sucesso: HTML Definitivo gerado em {html_path}.")
    except Exception as e:
        print(f"Erro processando {filename}: {e}")
