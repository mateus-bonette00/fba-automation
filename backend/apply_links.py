import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

files_to_fix = [
    "exports/Produtos_Daniel_20_02_2026_195500_CORRIGIDO.xlsx",
    "exports/Produtos_Daniel_20_02_2026_205843.xlsx",
    "exports/Produtos_Daniel_20_02_2026_212722.xlsx",
    "exports/Produtos_Daniel_20_02_2026_213201.xlsx"
]

base_dir = "/home/mateus/Documentos/Qota Store/códigos/fba-automation"

for filename in files_to_fix:
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        print(f"File not found: {path}")
        continue
        
    print(f"Processando {filename}...")
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        # Iterar sobre as células para achar os links
        count = 0
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val and isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val
                    # Manter legível em azul com sublinhado
                    cell.font = Font(color="0563C1", underline="single")
                    count += 1
                    
        wb.save(path)
        print(f"Sucesso: {count} links formatados em {filename}.")
    except Exception as e:
        print(f"Erro processando {filename}: {e}")
