import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def export_to_xlsx(accumulated_items, template_path, output_path):
    # Carrega o template mantendo formatação
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template não encontrado: {template_path}")
        
    wb = load_workbook(template_path)
    ws = wb.active

    # O template "SRAM 05_01_2026.xlsx" tem os headers na linha 1
    # Vamos descobrir as colunas com base no cabecalho
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column

    # Preenche os dados copiando o estilo da linha 2 (primeira linha de dados)
    # Assumimos que o cabecalho está na linha 1 e os dados comecam na 2
    
    start_row = 2
    # Extrair os estilos da linha 2 (row=2)
    style_row = 2
    for r_idx, item in enumerate(accumulated_items, start=start_row):
        
        # Mapear chaves do item para colunas (ajustado para o padrão Qota Store)
        mapping = {
            "produto": item.get('product_title', ''),
            "upc": item.get('upc', ''),
            "url fornecedor": item.get('url', ''),
            "amazon upc": f"https://www.amazon.com/s?k={item.get('upc', '')}" if item.get('upc') else "",
            "amazon titulo": f"https://www.amazon.com/s?k={item.get('product_title', '')}" if item.get('product_title') else "",
        }

        # Insere dados
        for header_name, col_idx in headers.items():
            val = ""
            for k, v in mapping.items():
                if k.lower() == header_name:
                    val = v
                    break
            
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.value = val
            
            # Copiar estilo da primeira linha de dados
            template_cell = ws.cell(row=style_row, column=col_idx)
            if template_cell.has_style:
                cell.font = copy(template_cell.font)
                cell.border = copy(template_cell.border)
                cell.fill = copy(template_cell.fill)
                cell.number_format = copy(template_cell.number_format)
                cell.protection = copy(template_cell.protection)
                cell.alignment = copy(template_cell.alignment)
                
            # Transforma explicitamente em fórmula HYPERLINK para ser 100% nativo e clicável no Google Planilhas
            if val and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = None # Remove metadados se houver
                cell.value = f'=HYPERLINK("{val}", "{val}")'
                # Colore de azul e sublinha para manter aspecto de link
                cell.font = Font(color="0563C1", underline="single")

    # Limpar as linhas extras do template original
    max_orig_row = ws.max_row
    if max_orig_row > max(start_row, r_idx):
        ws.delete_rows(r_idx + 1, max_orig_row - r_idx)

    # Ajusta largura das colunas e quebra de linha para garantir visibilidade
    for header_name, col_idx in headers.items():
        col_letter = get_column_letter(col_idx)
        if 'url' in header_name or 'amazon' in header_name:
            ws.column_dimensions[col_letter].width = 60
        elif 'produto' in header_name:
            ws.column_dimensions[col_letter].width = 80
        else:
            ws.column_dimensions[col_letter].width = 25
            
        for row in range(start_row, r_idx + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.alignment:
                cell.alignment = Alignment(wrap_text=True, horizontal=cell.alignment.horizontal, vertical='center')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

    wb.save(output_path)
    
    # GERADOR DE HTML PARALELO PARA CLICK INSTANTÂNEO NATIVO
    try:
        html_path = output_path.replace(".xlsx", ".html").replace("ARQUIVOS XLSX", "ARQUIVOS HTML")
        os.makedirs(os.path.dirname(html_path), exist_ok=True)
        html_content = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Produtos Capturados</title>",
            "<style>",
            ":root { --bg: #ffffff; --text: #333333; --table-bg: #f9f9f9; --border: #dddddd; --th-bg: #0563C1; --th-text: #ffffff; --link: #0066cc; --link-hover: #003399; --link-visited: #cc0000; --tr-hover: #f1f1f1; --clicked-bg: #ffe6e6; }",
            "@media (prefers-color-scheme: dark) {",
            "  :root { --bg: #121212; --text: #e0e0e0; --table-bg: #1e1e1e; --border: #333333; --th-bg: #0563C1; --th-text: #ffffff; --link: #4da6ff; --link-hover: #80bfff; --link-visited: #ff8080; --tr-hover: #2a2a2a; --clicked-bg: #4a1515; }",
            "}",
            "body { font-family: monospace; font-size: 14px; margin: 20px; background: var(--bg); color: var(--text); }",
            "table { border-collapse: collapse; width: 100%; margin-top: 10px; background: var(--table-bg); }",
            "th, td { border: 1px solid var(--border); padding: 8px; text-align: left; transition: background-color 0.3s; }",
            "th { background-color: var(--th-bg); color: var(--th-text); position: sticky; top: 0; }",
            "a { color: var(--link); text-decoration: none; display: block; width: 100%; height: 100%; }",
            "a:hover { text-decoration: underline; color: var(--link-hover); }",
            "a:visited { color: var(--link-visited); }",
            "tr:hover { background-color: var(--tr-hover); }",
            "td.clicked-cell { background-color: var(--clicked-bg) !important; }",
            "</style></head><body>",
            "<h2>Lote de Produtos Exportados</h2>",
            "<p>Dica: Segure <b>CTRL + Clique</b> nos links abaixo para abrir na hora sem sair dessa tela!</p>",
            "<table><thead><tr><th>Indice</th><th>Produto</th><th>UPC</th><th>URL Fornecedor</th><th>Amazon UPC</th><th>Amazon Título</th></tr></thead><tbody>"
        ]
        
        for index, item in enumerate(accumulated_items, start=1):
            produto = item.get('product_title', '')
            upc = item.get('upc', '')
            url = item.get('url', '')
            amz_upc = f"https://www.amazon.com/s?k={upc}" if upc else ""
            amz_tit = f"https://www.amazon.com/s?k={produto}" if produto else ""
            
            html_content.append("<tr>")
            html_content.append(f"<td>{index}</td>")
            html_content.append(f"<td>{produto}</td>")
            html_content.append(f"<td>{upc}</td>")
            html_content.append(f"<td><a href='{url}' target='_blank'>Abrir Fornecedor</a></td>" if url else "<td></td>")
            html_content.append(f"<td><a href='{amz_upc}' target='_blank'>Buscar UPC na Amazon</a></td>" if amz_upc else "<td></td>")
            html_content.append(f"<td><a href='{amz_tit}' target='_blank'>Buscar Titulo na Amazon</a></td>" if amz_tit else "<td></td>")
            html_content.append("</tr>")
            
        html_content.append("</tbody></table>")
        
        # Add JavaScript for LocalStorage click tracking
        js_script = """
        <script>
            document.addEventListener('DOMContentLoaded', () => {
                const links = document.querySelectorAll('td a');
                
                // Load clicked state from localStorage
                links.forEach(link => {
                    const url = link.getAttribute('href');
                    if(url && localStorage.getItem('clicked_' + url)) {
                        link.parentElement.classList.add('clicked-cell');
                    }
                    
                    // Add click listener
                    link.addEventListener('click', function() {
                        if(url) {
                            localStorage.setItem('clicked_' + url, 'true');
                            this.parentElement.classList.add('clicked-cell');
                        }
                    });
                });
            });
        </script>
        """
        html_content.append(js_script)
        html_content.append("</body></html>")
        
        with open(html_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html_content))
    except Exception as e:
        print(f"Erro ao gerar versão HTML: {e}")

    return output_path
