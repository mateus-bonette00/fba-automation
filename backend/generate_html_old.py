import sys
import os
from openpyxl import load_workbook

files_to_fix = [
    "exports/Produtos_Daniel_20_02_2026_195500_CORRIGIDO.xlsx",
    "exports/Produtos_Daniel_20_02_2026_205843.xlsx",
    "exports/Produtos_Daniel_20_02_2026_212722.xlsx",
    "exports/Produtos_Daniel_20_02_2026_213201.xlsx"
]

base_dir = "/home/mateus/Documentos/Qota Store/códigos/fba-automation"

for filename in files_to_fix:
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        print(f"File not found: {path}")
        continue
        
    print(f"Gerando HTML para {filename}...")
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        html_path = path.replace(".xlsx", ".html")
        html_content = [
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Produtos Capturados</title>",
            "<style>",
            "body { font-family: monospace; font-size: 14px; margin: 20px; background: #121212; color: #e0e0e0; }",
            "table { border-collapse: collapse; width: 100%; margin-top: 10px; background: #1e1e1e; }",
            "th, td { border: 1px solid #333; padding: 8px; text-align: left; }",
            "th { background-color: #0563C1; color: white; position: sticky; top: 0; }",
            "a { color: #4da6ff; text-decoration: none; }",
            "a:hover { text-decoration: underline; color: #80bfff; }",
            "tr:hover { background-color: #2a2a2a; }",
            "</style></head><body>",
            "<h2>Lote de Produtos Exportados</h2>",
            "<p>Dica: Segure <b>CTRL + Clique</b> nos links abaixo para abrir na hora sem sair dessa tela!</p>",
            "<table><thead><tr><th>Indice</th><th>Produto</th><th>UPC</th><th>URL Fornecedor</th><th>Amazon UPC</th><th>Amazon Título</th></tr></thead><tbody>"
        ]
        
        index = 1
        for row in range(2, ws.max_row + 1):
            produto = ws.cell(row=row, column=1).value or ""
            upc = ws.cell(row=row, column=2).value or ""
            
            # Fetch URLs (from hyperlinks if formula, or target if metadata)
            url = ""
            amz_upc = ""
            amz_tit = ""
            
            # Helper to extract url
            def get_url(cell):
                if cell.value and isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
                    try:
                        return cell.value.split('"')[1]
                    except: return ""
                if cell.hyperlink and cell.hyperlink.target:
                    return cell.hyperlink.target
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                    return cell.value
                return ""
            
            url = get_url(ws.cell(row=row, column=3))
            amz_upc = get_url(ws.cell(row=row, column=4))
            amz_tit = get_url(ws.cell(row=row, column=5))

            html_content.append("<tr>")
            html_content.append(f"<td>{index}</td>")
            html_content.append(f"<td>{produto}</td>")
            html_content.append(f"<td>{upc}</td>")
            html_content.append(f"<td><a href='{url}' target='_blank'>Abrir Fornecedor</a></td>" if url else "<td></td>")
            html_content.append(f"<td><a href='{amz_upc}' target='_blank'>Buscar UPC na Amazon</a></td>" if amz_upc else "<td></td>")
            html_content.append(f"<td><a href='{amz_tit}' target='_blank'>Buscar Titulo na Amazon</a></td>" if amz_tit else "<td></td>")
            html_content.append("</tr>")
            index += 1
            
        html_content.append("</tbody></table></body></html>")
        
        with open(html_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html_content))
            
        print(f"Sucesso: HTML gerado em {html_path}.")
    except Exception as e:
        print(f"Erro processando {filename}: {e}")
