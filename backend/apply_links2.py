import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

files_to_fix = [
    "exports/Produtos_Daniel_20_02_2026_195500_CORRIGIDO.xlsx",
    "exports/Produtos_Daniel_20_02_2026_205843.xlsx",
    "exports/Produtos_Daniel_20_02_2026_212722.xlsx",
    "exports/Produtos_Daniel_20_02_2026_213201.xlsx"
]

base_dir = "/home/mateus/Documentos/Qota Store/códigos/fba-automation"

for filename in files_to_fix:
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        print(f"File not found: {path}")
        continue
        
    print(f"Processando {filename} com HYPERLINK formula...")
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        count = 0
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                # Pode já ser uma fórmula antiga se o cara rodar o script duas vezes, ou ser metadata
                if val and isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = None
                    cell.value = f'=HYPERLINK("{val}", "{val}")'
                    cell.font = Font(color="0563C1", underline="single")
                    count += 1
                elif val and isinstance(val, str) and val.startswith("=HYPERLINK"):
                    # Ja eh hyperlink
                    pass
                elif cell.hyperlink and cell.hyperlink.target:
                    # Tinha hyperlink em metadata (do script anterior)
                    t_val = cell.hyperlink.target
                    cell.hyperlink = None
                    cell.value = f'=HYPERLINK("{t_val}", "{t_val}")'
                    cell.font = Font(color="0563C1", underline="single")
                    count += 1
                    
        wb.save(path)
        print(f"Sucesso: {count} links convertidos para formula em {filename}.")
    except Exception as e:
        print(f"Erro processando {filename}: {e}")
